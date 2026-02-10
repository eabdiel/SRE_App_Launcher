from __future__ import annotations

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from models import AutomationProject


STEPS_HEADERS = [
    "Step#",
    "Channel",
    "Action",
    "WindowTitle",
    "ProcessName",
    "PID",
    "HWND",
    "LocatorType",
    "Locator",
    "InputRef",
    "Value",
    "OutputRef",
    "WaitMs",
    "Notes",
]

DATA_HEADERS = [
    "Key",
    "Value",
    "Type",
    "PromptOnRun",
    "DefaultValue",
]


def export_project_xlsx(project: AutomationProject, path: str) -> None:
    wb = Workbook()

    ws_steps = wb.active
    ws_steps.title = "Steps"
    ws_steps.append(STEPS_HEADERS)

    for idx, s in enumerate(project.steps, start=1):
        ws_steps.append([
            idx,
            s.channel,
            s.action,
            s.window_title,
            s.process_name,
            s.pid or "",
            s.hwnd or "",
            s.locator_type,
            s.locator,
            s.input_ref,
            s.value,
            s.output_ref,
            s.wait_ms,
            s.notes,
        ])

    ws_data = wb.create_sheet("Data")
    ws_data.append(DATA_HEADERS)
    for d in project.data:
        ws_data.append([d.key, d.value, d.type, "Y" if d.prompt_on_run else "N", d.default_value])

    _autosize(ws_steps)
    _autosize(ws_data)

    wb.save(path)


def _autosize(ws) -> None:
    for col in range(1, ws.max_column + 1):
        max_len = 0
        for row in range(1, ws.max_row + 1):
            v = ws.cell(row=row, column=col).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 2, 60)
